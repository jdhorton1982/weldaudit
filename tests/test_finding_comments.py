"""Comments an auditor writes on a finding.

The point of the feature is that they outlive the audit. Every run deletes
this project's findings and builds them again, so a comment stored on a
finding's id lasts until the next run — and the next run is exactly when
somebody wants to see what they said last time.

They are therefore keyed on what identifies the *problem* — rule, segment,
subject — rather than the row. The message is deliberately not part of the
key: improving a rule rewrites it, and the comment is still about the same
heat on the same line.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pytest  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

from weldaudit.api import create_app  # noqa: E402
from weldaudit.db import Database  # noqa: E402
from weldaudit.report import write_csv, write_excel  # noqa: E402


def _finding(db, pid, rule="MTR-02", segment="16 PW", subject="Norvale",
             message="not on the approved list", status="open"):
    with db.tx() as c:
        c.execute(
            """INSERT INTO finding(project_id, run_id, rule, severity, segment,
                                   subject, message, status)
               VALUES(?, 'r1', ?, 'critical', ?, ?, ?, ?)""",
            (pid, rule, segment, subject, message, status))
        return c.execute("SELECT last_insert_rowid() i").fetchone()[0]


@pytest.fixture
def job(tmp_path):
    db = Database(tmp_path / "t.db")
    root = tmp_path / "Job"
    root.mkdir()
    return db, db.upsert_project("Job", root)


@pytest.fixture
def api(job, tmp_path):
    return TestClient(create_app(tmp_path / "t.db"))


# -- writing one -------------------------------------------------------------

def test_a_comment_is_stored_and_shows_at_once(api, job):
    """Without waiting for another audit — the finding on screen carries it."""
    db, pid = job
    fid = _finding(db, pid)
    r = api.post(f"/api/findings/{fid}/comment",
                 json={"text": "Rang the mill, cert is coming Monday"})
    assert r.status_code == 200
    assert r.json()["comment"] == "Rang the mill, cert is coming Monday"
    assert db.one("SELECT note FROM finding WHERE id=?", (fid,))["note"] \
        == "Rang the mill, cert is coming Monday"


def test_writing_again_replaces_it(api, job):
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/comment", json={"text": "first"})
    api.post(f"/api/findings/{fid}/comment", json={"text": "second"})
    assert len(db.comments(pid)) == 1
    assert db.comments(pid)[0]["comment"] == "second"


def test_an_empty_comment_clears_it(api, job):
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/comment", json={"text": "never mind"})
    api.post(f"/api/findings/{fid}/comment", json={"text": "   "})
    assert db.comments(pid) == []
    assert db.one("SELECT note FROM finding WHERE id=?", (fid,))["note"] is None


def test_whitespace_is_tidied(api, job):
    db, pid = job
    fid = _finding(db, pid)
    got = api.post(f"/api/findings/{fid}/comment",
                   json={"text": "  chased   the   mill \n"}).json()
    assert got["comment"] == "chased the mill"


def test_commenting_on_a_finding_that_does_not_exist(api):
    assert api.post("/api/findings/999999/comment",
                    json={"text": "x"}).status_code == 404


# -- surviving the audit that rebuilds the findings --------------------------

def test_a_comment_survives_the_findings_being_rebuilt(api, job):
    """The whole point. `clear_findings` empties the table on every run."""
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/comment", json={"text": "cert on order"})

    db.clear_findings(pid)                     # what an audit does
    again = _finding(db, pid, message="rewritten by an improved rule")
    assert db.one("SELECT note FROM finding WHERE id=?", (again,))["note"] is None

    assert db.reattach_comments(pid) == 1
    assert db.one("SELECT note FROM finding WHERE id=?",
                  (again,))["note"] == "cert on order"


def test_it_reattaches_to_the_same_problem_not_the_same_row(job):
    """Keyed on rule, segment and subject. A finding whose message changed is
    the same finding; one about a different heat is not."""
    db, pid = job
    fid = _finding(db, pid, subject="Norvale")
    db.set_comment(pid, "MTR-02", "16 PW", "Norvale", "spoke to QA")
    db.clear_findings(pid)

    same = _finding(db, pid, subject="Norvale", message="quite different wording")
    other = _finding(db, pid, subject="Halden")
    db.reattach_comments(pid)
    assert db.one("SELECT note FROM finding WHERE id=?", (same,))["note"] == "spoke to QA"
    assert db.one("SELECT note FROM finding WHERE id=?", (other,))["note"] is None


def test_a_comment_outlives_a_finding_that_stops_being_raised(api, job):
    """If the problem is fixed the finding goes, and the comment stays on
    record rather than disappearing with it."""
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/comment", json={"text": "resolved with the mill"})
    db.clear_findings(pid)
    assert [c["comment"] for c in db.comments(pid)] == ["resolved with the mill"]
    assert api.get("/api/comments", params={"project_id": pid}).json()[0]["comment"] \
        == "resolved with the mill"


def test_comments_belong_to_one_job(job, tmp_path):
    db, pid = job
    other_root = tmp_path / "Other"
    other_root.mkdir()
    other = db.upsert_project("Other", other_root)
    db.set_comment(pid, "MTR-02", "16 PW", "Norvale", "mine")
    assert db.comments(other) == []


# -- the exports -------------------------------------------------------------

def test_the_csv_carries_a_comments_column(api, job, tmp_path):
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/comment", json={"text": "chased Monday"})
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})

    import csv
    out = write_csv(db, pid, tmp_path / "x.csv")
    with out.open(encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    assert rows[0][-1] == "Comments"
    assert rows[1][-1] == "chased Monday"


def test_the_excel_carries_a_comments_column(api, job, tmp_path):
    import openpyxl

    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/comment", json={"text": "chased Monday"})
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})

    out = write_excel(db, pid, tmp_path / "x.xlsx")
    ws = openpyxl.load_workbook(out)["Findings"]
    headers = [c.value for c in ws[1]]
    assert headers[-1] == "Comments"
    assert ws.cell(row=2, column=len(headers)).value == "chased Monday"


def test_an_uncommented_finding_exports_an_empty_cell(job, tmp_path):
    db, pid = job
    _finding(db, pid, status="accepted")
    import csv
    out = write_csv(db, pid, tmp_path / "x.csv")
    with out.open(encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    assert rows[1][-1] == ""


# -- the decision survives too -----------------------------------------------

def test_a_decision_survives_the_findings_being_rebuilt(api, job):
    """The same problem as the comment, and the one that mattered more: a
    morning spent accepting and dismissing was gone by the afternoon's run,
    and every finding came back open with nothing to say it had been read."""
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})

    db.clear_findings(pid)
    again = _finding(db, pid, message="rewritten by an improved rule")
    assert db.one("SELECT status FROM finding WHERE id=?", (again,))["status"] == "open"

    db.reattach_comments(pid)
    assert db.one("SELECT status FROM finding WHERE id=?",
                  (again,))["status"] == "accepted"


def test_reopening_clears_the_stored_decision(api, job):
    """Back on the list is not a decision, so nothing should be remembered
    that would put it back into accepted on the next run."""
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})
    api.post(f"/api/findings/{fid}/status", json={"status": "open"})

    db.clear_findings(pid)
    again = _finding(db, pid)
    db.reattach_comments(pid)
    assert db.one("SELECT status FROM finding WHERE id=?",
                  (again,))["status"] == "open"


def test_a_comment_and_a_decision_share_one_row(api, job):
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/comment", json={"text": "cert on order"})
    api.post(f"/api/findings/{fid}/status", json={"status": "dismissed"})
    rows = db.comments(pid)
    assert len(rows) == 1
    assert rows[0]["comment"] == "cert on order"
    assert rows[0]["status"] == "dismissed"


def test_reopening_does_not_take_the_comment_with_it(api, job):
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/comment", json={"text": "spoke to the mill"})
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})
    api.post(f"/api/findings/{fid}/status", json={"status": "open"})
    assert db.one("SELECT note FROM finding WHERE id=?",
                  (fid,))["note"] == "spoke to the mill"
    assert db.comments(pid)[0]["status"] is None


def test_clearing_the_comment_leaves_the_decision(api, job):
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})
    api.post(f"/api/findings/{fid}/comment", json={"text": "note"})
    api.post(f"/api/findings/{fid}/comment", json={"text": ""})
    assert db.comments(pid)[0]["status"] == "accepted"
    assert db.one("SELECT status FROM finding WHERE id=?",
                  (fid,))["status"] == "accepted"


# -- what goes out, and what does not ----------------------------------------
#
# "can exceptions that are no issue be excluded from download. they dont need
# to be send to the person revising documents." The report is a punch list for
# somebody else to work through; an item the auditor has already checked and
# cleared is not work for them, and every row of it costs the ones that matter
# some attention.


def _read_csv(path):
    import csv
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.reader(fh))


def test_a_no_issue_finding_is_left_out_of_the_csv(api, job, tmp_path):
    db, pid = job
    _finding(db, pid, rule="MTR-02", subject="Kandal", status="accepted")
    drop = _finding(db, pid, rule="AB-07", subject="42+72")
    api.post(f"/api/findings/{drop}/status", json={"status": "dismissed"})

    rows = _read_csv(write_csv(db, pid, tmp_path / "x.csv"))
    subjects = [r[3] for r in rows[1:]]
    assert subjects == ["Kandal"]


def test_a_no_issue_finding_is_left_out_of_the_excel(api, job, tmp_path):
    import openpyxl

    db, pid = job
    _finding(db, pid, rule="MTR-02", subject="Kandal", status="accepted")
    drop = _finding(db, pid, rule="AB-07", subject="42+72")
    api.post(f"/api/findings/{drop}/status", json={"status": "dismissed"})

    ws = openpyxl.load_workbook(write_excel(db, pid, tmp_path / "x.xlsx"))["Findings"]
    assert [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)] == ["Kandal"]


def test_a_confirmed_issue_still_goes_out(api, job, tmp_path):
    """Marking something a real exception must not remove it — that is the
    half of the pair that most needs to reach the contractor."""
    db, pid = job
    fid = _finding(db, pid, subject="Kandal")
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})
    rows = _read_csv(write_csv(db, pid, tmp_path / "x.csv"))
    assert [r[3] for r in rows[1:]] == ["Kandal"]
    assert rows[1][8] == "ISSUE"


def test_an_unreviewed_finding_is_held_back(job, tmp_path):
    """Only findings an auditor has confirmed go to the contractor. The cost
    is a report that is empty until the list has been worked through, which is
    why the window says how many are being held back before it writes."""
    db, pid = job
    _finding(db, pid, subject="Kandal")
    assert _read_csv(write_csv(db, pid, tmp_path / "x.csv"))[1:] == []


def test_marking_it_an_issue_puts_it_in_the_report(api, job, tmp_path):
    db, pid = job
    fid = _finding(db, pid, subject="Kandal")
    api.post(f"/api/findings/{fid}/status", json={"status": "dismissed"})
    assert _read_csv(write_csv(db, pid, tmp_path / "a.csv"))[1:] == []
    api.post(f"/api/findings/{fid}/status", json={"status": "open"})
    assert _read_csv(write_csv(db, pid, tmp_path / "b.csv"))[1:] == [],         "back on the list is not a verdict, so still not in the report"
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})
    assert [r[3] for r in _read_csv(write_csv(db, pid, tmp_path / "c.csv"))[1:]] == ["Kandal"]


def test_it_is_left_out_of_the_file_not_thrown_away(api, job):
    """The auditor has to be able to show they checked it."""
    db, pid = job
    fid = _finding(db, pid, subject="Kandal")
    api.post(f"/api/findings/{fid}/comment", json={"text": "rang the mill, cert is fine"})
    api.post(f"/api/findings/{fid}/status", json={"status": "dismissed"})

    still = api.get("/api/findings", params={"project_id": pid, "status": "dismissed"}).json()
    assert [r["subject"] for r in still["rows"]] == ["Kandal"]
    assert db.comments(pid)[0]["comment"] == "rang the mill, cert is fine"


def test_the_decision_survives_a_re_audit_and_stays_out(api, job, tmp_path):
    """The case that matters: cleared in the morning, re-audited, and the
    report that goes out in the afternoon must not have it back."""
    db, pid = job
    fid = _finding(db, pid, rule="MTR-02", subject="Kandal")
    api.post(f"/api/findings/{fid}/status", json={"status": "dismissed"})

    db.clear_findings(pid)                     # what an audit does
    _finding(db, pid, rule="MTR-02", subject="Kandal", message="reworded rule")
    db.reattach_comments(pid)

    assert _read_csv(write_csv(db, pid, tmp_path / "x.csv"))[1:] == []


# -- the words in the Status column ------------------------------------------
#
# The database has called a confirmed exception "accepted" since before the
# buttons said Issue, and renaming the stored value would discard every
# decision already taken. So the report translates at the edge, where it is
# written for somebody who never saw the app.


def test_the_status_column_says_ISSUE(api, job, tmp_path):
    db, pid = job
    fid = _finding(db, pid, subject="Kandal")
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})
    rows = _read_csv(write_csv(db, pid, tmp_path / "x.csv"))
    assert rows[0][8] == "Status"
    assert rows[1][8] == "ISSUE"


def test_the_excel_status_column_says_ISSUE(api, job, tmp_path):
    import openpyxl

    db, pid = job
    fid = _finding(db, pid, subject="Kandal")
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})
    ws = openpyxl.load_workbook(write_excel(db, pid, tmp_path / "x.xlsx"))["Findings"]
    assert ws.cell(row=1, column=9).value == "Status"
    assert ws.cell(row=2, column=9).value == "ISSUE"


def test_the_stored_value_is_untouched(api, job):
    """Translated for the report only. Renaming it in the database would
    throw away every decision already taken."""
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})
    assert db.one("SELECT status FROM finding WHERE id=?", (fid,))["status"] == "accepted"


# -- saying what is being held back ------------------------------------------
#
# Only Issues go out, so a report taken before the list has been worked
# through is empty -- and an empty report reads exactly like a clean package.
# The window asks before it writes, using these counts.


def test_the_scope_counts_what_goes_and_what_stays(api, job):
    db, pid = job
    a = _finding(db, pid, rule="MTR-02", subject="Kandal")
    b = _finding(db, pid, rule="AB-07", subject="42+72")
    _finding(db, pid, rule="NDE-01", subject="W-14")        # left unreviewed
    api.post(f"/api/findings/{a}/status", json={"status": "accepted"})
    api.post(f"/api/findings/{b}/status", json={"status": "dismissed"})

    s = api.get("/api/report-scope", params={"project_id": pid}).json()
    assert s == {"going_out": 1, "unreviewed": 1, "no_issue": 1}


def test_an_untouched_job_reports_an_empty_download(api, job):
    """The dangerous case, and the reason the prompt exists."""
    db, pid = job
    for n in range(5):
        _finding(db, pid, rule=f"MTR-0{n}", subject=f"s{n}")
    s = api.get("/api/report-scope", params={"project_id": pid}).json()
    assert s["going_out"] == 0
    assert s["unreviewed"] == 5


def test_a_fully_worked_job_has_nothing_held_back(api, job):
    db, pid = job
    fid = _finding(db, pid)
    api.post(f"/api/findings/{fid}/status", json={"status": "accepted"})
    s = api.get("/api/report-scope", params={"project_id": pid}).json()
    assert s["unreviewed"] == 0
    assert s["going_out"] == 1
