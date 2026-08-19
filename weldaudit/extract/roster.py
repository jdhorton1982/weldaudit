"""The contractor's project welder log — who was on the job, and when.

One template across both jobs that keep one, header on row 6, sixteen
spreadsheets and ninety-eight welders.  It carries what the certificates
cannot: the welder's **name** against the stencil, the date they arrived and
left, and the contractor's own stated requalification due date.

That last one matters most.  WLD-03 has to infer continuity from a gap of more
than 183 days between welds, because nothing else on the job says when a ticket
lapses.  The roster states it outright.
"""

from __future__ import annotations

import re
import warnings
from datetime import date, datetime
from typing import Any

import openpyxl

from ..db import Database

warnings.filterwarnings("ignore", module="openpyxl")

SOURCE = "welder_roster_xlsx"

#: Column labels as the template writes them, matched as a prefix on the
#: normalised header cell.
_COLUMNS: tuple[tuple[str, str], ...] = (
    ("name", "welder name"),
    ("stencil", "welder stencil"),
    ("material", "cert for"),
    ("cert_date", "cert test date"),
    ("requal_date", "requal date"),
    ("next_requal", "next requal"),
    ("arrived", "date arrived"),
    ("left_job", "date left"),
    ("reason", "reason no longer"),
)

#: The rightmost column of the sheet is a block of standing instructions to the
#: contractor ("***If you receive a cert from elsewhere, please upload..."),
#: repeated down every row. It is not data.
_INSTRUCTIONS = re.compile(r"^\s*\*{2,}|please upload|so that it can be", re.IGNORECASE)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


#: `19/9/25` appears once where every other cell is a real date. Day first,
#: because the other two orderings would make it the 19th month.
_SLASHED = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$")


def _iso(value: Any) -> str:
    """A date cell, whether Excel stored it as a date or the crew typed text."""
    text = _clean(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    if m := _SLASHED.match(text):
        day, month, year = (int(g) for g in m.groups())
        if day <= 12 < month:                # unambiguous the other way round
            day, month = month, day
        if year < 100:
            year += 2000
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return ""
    for fmt in ("%m-%d-%y", "%m-%d-%Y", "%m.%d.%y", "%m.%d.%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _locate(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """Header row index and ``{field: column}``, found by label."""
    for i, row in enumerate(rows[:20]):
        labels = {j: _clean(v).lower() for j, v in enumerate(row) if v is not None}
        if not any(text.startswith("welder name") for text in labels.values()):
            continue
        found: dict[str, int] = {}
        for field, prefix in _COLUMNS:
            for j, text in labels.items():
                if text.startswith(prefix) and field not in found:
                    found[field] = j
        return i, found
    return None


def _parse(path: str) -> list[dict]:
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: list[dict] = []
    for sheet in book.sheetnames:
        rows = list(book[sheet].iter_rows(values_only=True))
        located = _locate(rows)
        if not located:
            continue
        header_row, columns = located

        cell = lambda row, field: (                       # noqa: E731
            row[columns[field]]
            if field in columns and columns[field] < len(row) else None)

        for offset, row in enumerate(rows[header_row + 1:], start=1):
            name = _clean(cell(row, "name"))
            stencil = _clean(cell(row, "stencil")).upper()
            if not name or _INSTRUCTIONS.match(name):
                continue
            out.append({
                "row_no": offset, "name": name, "stencil": stencil,
                "material": _clean(cell(row, "material")).upper(),
                "cert_date": _iso(cell(row, "cert_date")),
                "requal_date": _iso(cell(row, "requal_date")),
                "next_requal": _iso(cell(row, "next_requal")),
                "arrived": _iso(cell(row, "arrived")),
                "left_job": _iso(cell(row, "left_job")),
                "reason": _clean(cell(row, "reason")),
            })
    return out


def extract(db: Database, project_id: int) -> int:
    """Load every project welder log. Returns the welder rows recorded."""
    documents = db.q(
        """SELECT id, path, filename, segment, fingerprint FROM document
           WHERE project_id=? AND kind='welder_roster'
             AND ext IN ('.xlsx','.xlsm','.xls')""",
        (project_id,),
    )

    records: list[tuple] = []
    seen: set[str] = set()
    for doc in documents:
        key = doc["fingerprint"] or f"doc:{doc['id']}"
        if key in seen:
            continue
        seen.add(key)
        try:
            rows = _parse(doc["path"])
        except Exception:                     # noqa: BLE001 - a bad workbook is
            continue                          # not worth failing the run over
        for r in rows:
            records.append((
                project_id, doc["id"], doc["fingerprint"], doc["segment"] or "",
                r["row_no"], r["name"], r["stencil"], r["material"],
                r["cert_date"], r["requal_date"], r["next_requal"],
                r["arrived"], r["left_job"], r["reason"], SOURCE,
            ))

    with db.tx() as c:
        c.execute("DELETE FROM welder_roster WHERE project_id=? AND source=?",
                  (project_id, SOURCE))
        if records:
            c.executemany(
                """INSERT INTO welder_roster
                   (project_id, document_id, fingerprint, segment, row_no, name,
                    stencil, material, cert_date, requal_date, next_requal,
                    arrived, left_job, reason, source)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                records,
            )
    return len(records)
