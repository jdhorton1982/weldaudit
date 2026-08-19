"""Flange logs and flange maps.

Three things come out of the flange section, and all three are free — no page
has to be sent to a model:

``flange``
    Bluewater's 27 torque logs are .xlsx on one template, 587 bolted joints with
    size, class, gasket, bolt count, four torque rounds, the wrench serial and
    the inspector's sign-off.  This is the richest deterministic table in the
    whole audit.

``flange_map``
    The maps are drawings, but their balloon numbers survive into the PDF text
    layer as a bare run of integers.  The count of that run is how many flanges
    the drawing shows, which is what the log has to account for.

``instrument_cal``
    Torque wrench certificates.  Bluewater names most of them for nothing but the
    serial (`0322600192.pdf`), which is only safe to read as a serial because
    the document is already known to be filed in the flange section.

The template puts data in every other column, and the header on row 8 (1-based)
of every workbook in the corpus.  Rather than trust that, the header row is
found by looking for the "Flange #" cell and the columns are located by their
own labels, so a log with a shifted layout still parses.
"""

from __future__ import annotations

import re
import warnings
from datetime import date, datetime
from typing import Any

import openpyxl
import pymupdf

from ..db import Database
from ..instruments import parse_bare_serial, serial_key

warnings.filterwarnings("ignore", module="openpyxl")

SOURCE = "flange_log_xlsx"
MAP_SOURCE = "flange_map_pdf"
WRENCH_SOURCE = "flange_wrench_filename"

#: Column labels as the template writes them, mapped to our field names. The
#: match is a prefix on the normalised label, so trailing spaces and the long
#: "Final Round Clockwise or Counterclockwise Pattern" both land.
_COLUMNS: tuple[tuple[str, str], ...] = (
    ("flange_no", "flange #"),
    ("nps", "size"),
    ("pressure_class", "class"),
    ("gasket", "gasket type"),
    ("bolts", "bolts used"),
    ("lubricant", "kopr-kote"),
    ("round1", "round 1"),
    ("round2", "round 2"),
    ("round3", "round 3"),
    ("round4", "round 4"),
    ("pattern", "final round"),
    ("wrench", "torque wrench serial"),
    ("cert_checked", "copy of torque wrench"),
    ("inspector", "inspector initials"),
    ("notes", "notes"),
)

#: Header cells above the table. The job start date is worth having: it is the
#: only thing on the job that says a bolt-up dated 2006 is a typo.
_HEADER_FIELDS = (("line_size", "line size"), ("service", "service"),
                  ("job_start", "job start"))


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.]", "", _clean(value))
    try:
        return float(text) if text else None
    except ValueError:
        return None


_DATE_IN_TEXT = re.compile(r"\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4}|\d{4}-\d{2}-\d{2}")


def split_signoff(value: Any) -> tuple[str, str]:
    """Initials and date out of one "Inspector Initials & Date" cell.

    The column holds `WL 9/26/25`, or bare initials, or - where the crew typed
    into a date-formatted cell - a datetime with no initials at all.
    """
    text = _clean(value)
    if not text:
        return "", ""
    when = ""
    if m := _DATE_IN_TEXT.search(text):
        when = _iso(m.group(0))
        text = (text[:m.start()] + " " + text[m.end():]).strip(" -,")
    initials = re.sub(r"[^A-Za-z./]", "", text).strip(".")
    return initials.upper(), when


def _iso(text: str) -> str:
    for fmt in ("%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y", "%m-%d-%y", "%m-%d-%Y",
                "%m.%d.%y", "%m.%d.%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


# ---------------------------------------------------------------------------
# Flange logs (.xlsx)
# ---------------------------------------------------------------------------


def _locate(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """The header row index and ``{field: column}``, found by label."""
    for i, row in enumerate(rows[:20]):
        labels = {j: _clean(v).lower() for j, v in enumerate(row) if v is not None}
        if not any(text.startswith("flange #") for text in labels.values()):
            continue
        found: dict[str, int] = {}
        for field, prefix in _COLUMNS:
            for j, text in labels.items():
                if text.startswith(prefix) and field not in found:
                    found[field] = j
        return i, found
    return None


def _header_values(rows: list[tuple], upto: int) -> dict[str, str]:
    """Line size and service, read from the block above the table."""
    out: dict[str, str] = {}
    for row in rows[:upto]:
        cells = [(j, _clean(v)) for j, v in enumerate(row) if v is not None]
        for field, label in _HEADER_FIELDS:
            for k, (j, text) in enumerate(cells):
                if text.lower().startswith(label) and k + 1 < len(cells):
                    out.setdefault(field, cells[k + 1][1])
    return out


def _parse_workbook(path: str) -> list[dict]:
    """Every flange row in every sheet of one log."""
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: list[dict] = []
    for name in book.sheetnames:
        rows = list(book[name].iter_rows(values_only=True))
        located = _locate(rows)
        if not located:
            continue
        header_row, columns = located
        header = _header_values(rows, header_row)

        cell = lambda row, field: (                       # noqa: E731
            row[columns[field]]
            if field in columns and columns[field] < len(row) else None)

        for offset, row in enumerate(rows[header_row + 1:], start=1):
            flange_no = _clean(cell(row, "flange_no"))
            if not flange_no and _number(cell(row, "nps")) is None:
                continue
            initials, when = split_signoff(cell(row, "inspector"))
            wrench = _clean(cell(row, "wrench"))
            out.append({
                "sheet": name, "row_no": offset, "flange_no": flange_no,
                "nps": _number(cell(row, "nps")),
                "pressure_class": _number(cell(row, "pressure_class")),
                "gasket": _clean(cell(row, "gasket")).upper(),
                "bolts": _number(cell(row, "bolts")),
                "lubricant": _clean(cell(row, "lubricant")).upper(),
                "round1": _number(cell(row, "round1")),
                "round2": _number(cell(row, "round2")),
                "round3": _number(cell(row, "round3")),
                "round4": _number(cell(row, "round4")),
                "pattern": _clean(cell(row, "pattern")).upper(),
                "wrench": wrench, "wrench_key": serial_key(wrench),
                "cert_checked": _clean(cell(row, "cert_checked")).upper(),
                "inspector": initials, "bolted_on": when,
                "notes": (notes := _clean(cell(row, "notes"))),
                # The NOTES column mostly holds the isometric this joint is
                # on, but some crews put a date there instead, so only a
                # drawing-shaped value is recorded as one.
                "drawing_no": (m.group(0).upper()
                               if (m := _DRAWING.search(notes)) else ""),
                "line_size": header.get("line_size", ""),
                "service": header.get("service", ""),
                "job_start": _iso(header.get("job_start", "")[:10])
                             or header.get("job_start", "")[:10],
            })
    return out


# ---------------------------------------------------------------------------
# Flange maps (.pdf)
# ---------------------------------------------------------------------------

#: A drawing number as the maps and logs write them: `6-600-8-XTO-FG-0502`,
#: `16-B60-PW-0401-01`, `8-B1-0-PO-0122`, `DTD22MP-FG-4-2C`.
_DRAWING = re.compile(
    r"\b\d{1,2}[\"']?[-_][A-Z0-9]{2,4}(?:[-_][A-Z0-9]{1,6}){2,5}\b|"
    r"\bDTD22MP[-_][A-Z]{2}[-_]\d{1,2}[-_][0-9A-Z]+\b",
    re.IGNORECASE,
)


def balloon_count(numbers: list[int]) -> int:
    """How many flanges a drawing balloons.

    Flange maps number their balloons from 1 upwards, so the answer is the
    longest run starting at 1.  Taking the maximum instead would read the
    ``45`` and ``90`` written against elbows as balloon numbers and claim a
    13-flange drawing has ninety.
    """
    present = set(numbers)
    n = 0
    while n + 1 in present:
        n += 1
    return n


#: The same template, printed to PDF instead of saved as a workbook. Kestrel 8
#: files its logs this way, and the text layer's reading order is scrambled -
#: values come out grouped by column, not by row - so the rows have to be
#: rebuilt from the word coordinates. These are the header words that identify
#: each column; the x centre of the match becomes that column's anchor.
_PDF_COLUMNS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("flange_no", ("flange",)),
    ("nps", ("size",)),
    ("pressure_class", ("class",)),
    ("gasket", ("gasket",)),
    ("bolts", ("bolts",)),
    ("lubricant", ("kopr-kote",)),
    ("round1", ("30%",)),
    ("round2", ("60%",)),
    ("round3", ("100%",)),
    ("pattern", ("pattern",)),
    ("wrench", ("serial",)),
    ("cert_checked", ("(verify",)),
    ("inspector", ("initials",)),
    ("notes", ("notes",)),
)

#: A flange number: the leftmost column, a bare small integer. Anchoring rows
#: on it is what keeps a value that wrapped onto its own y-band attached to
#: the joint above rather than becoming a row of its own.
_FLANGE_NO = re.compile(r"^\d{1,3}$")


def _columns_from_header(words: list) -> dict[str, float] | None:
    """``{field: x centre}`` from the printed header, or None if not a log."""
    anchors: dict[str, float] = {}
    for x0, _y0, x1, _y1, text, *_ in words:
        lowered = text.strip().lower()
        for field, labels in _PDF_COLUMNS:
            if field not in anchors and lowered in labels:
                anchors[field] = (x0 + x1) / 2
    return anchors if {"flange_no", "round3", "wrench"} <= set(anchors) else None


def _assign(words: list, anchors: dict[str, float]) -> dict[str, str]:
    """Each word to its nearest column, joined in reading order."""
    buckets: dict[str, list[tuple[float, str]]] = {}
    for x0, _y0, x1, _y1, text, *_ in words:
        centre = (x0 + x1) / 2
        field = min(anchors, key=lambda f: abs(anchors[f] - centre))
        # A word more than half a column away from every anchor is stray ink
        # from the printed boilerplate, not a value.
        if abs(anchors[field] - centre) > 40:
            continue
        buckets.setdefault(field, []).append((x0, text))
    return {f: " ".join(t for _x, t in sorted(v)) for f, v in buckets.items()}


def _parse_pdf_log(path: str) -> list[dict]:
    """Rebuild a printed flange log's rows from word coordinates."""
    out: list[dict] = []
    with pymupdf.open(path) as doc:
        for page in doc:
            words = page.get_text("words")
            anchors = _columns_from_header(words)
            if not anchors:
                continue
            header_y = max(
                (y1 for _x0, _y0, x1, y1, text, *_ in words
                 if text.strip().lower() in ("flange", "notes")), default=0)

            # Row anchors: the flange numbers running down the left column.
            starts = sorted(
                (y0, text) for x0, y0, x1, y1, text, *_ in words
                if y0 > header_y and _FLANGE_NO.match(text.strip())
                and abs((x0 + x1) / 2 - anchors["flange_no"]) <= 25
            )
            if not starts:
                continue

            for i, (top, flange_no) in enumerate(starts):
                bottom = starts[i + 1][0] if i + 1 < len(starts) else 1e9
                band = [w for w in words if top - 3 <= w[1] < bottom - 3]
                fields = _assign(band, anchors)
                initials, when = split_signoff(fields.get("inspector", ""))
                wrench = fields.get("wrench", "").strip()
                notes = fields.get("notes", "").strip()
                out.append({
                    "sheet": f"page {page.number + 1}", "row_no": i + 1,
                    "flange_no": flange_no.strip(),
                    "nps": _number(fields.get("nps")),
                    "pressure_class": _number(fields.get("pressure_class")),
                    "gasket": fields.get("gasket", "").strip().upper(),
                    "bolts": _number(fields.get("bolts")),
                    "lubricant": fields.get("lubricant", "").strip().upper(),
                    "round1": _number(fields.get("round1")),
                    "round2": _number(fields.get("round2")),
                    "round3": _number(fields.get("round3")),
                    # The printed form has no separate "around the world"
                    # column, so there is no fourth round to check.
                    "round4": None,
                    "pattern": fields.get("pattern", "").strip().upper(),
                    "wrench": wrench, "wrench_key": serial_key(wrench),
                    "cert_checked": fields.get("cert_checked", "").strip().upper(),
                    "inspector": initials, "bolted_on": when, "notes": notes,
                    "drawing_no": (m.group(0).upper()
                                   if (m := _DRAWING.search(notes)) else ""),
                    "line_size": "", "service": "", "job_start": "",
                })
    return out


def _parse_map(path: str) -> tuple[int, list[str]]:
    with pymupdf.open(path) as doc:
        text = " ".join(page.get_text() for page in doc)
    numbers = [int(x) for x in re.findall(r"\b(\d{1,3})\b", text)]
    drawings = sorted({m.group(0).upper() for m in _DRAWING.finditer(text)})
    return balloon_count(numbers), drawings


# ---------------------------------------------------------------------------


def extract(db: Database, project_id: int) -> tuple[int, int, int]:
    """Load flange logs, flange maps and torque wrench certificates.

    Returns ``(flange rows, maps read, wrench certificates)``.
    """
    documents = db.q(
        """SELECT id, path, filename, ext, segment, fingerprint FROM document
           WHERE project_id=? AND kind='flange_map'""",
        (project_id,),
    )

    flanges: list[tuple] = []
    maps: list[tuple] = []
    wrenches: list[tuple] = []
    seen: set[str] = set()

    for doc in documents:
        # The same log is filed into several segment books, as everything else
        # on these jobs is; one physical log is one set of joints.
        key = doc["fingerprint"] or f"doc:{doc['id']}"
        if key in seen:
            continue
        seen.add(key)

        rows: list[dict] = []
        if (doc["ext"] or "").lower() == ".xlsx":
            try:
                rows = _parse_workbook(doc["path"])
            except Exception:                    # noqa: BLE001 - a bad workbook
                rows = []                        # is not worth failing the run
        elif not parse_bare_serial(doc["filename"]).serial:
            # A PDF that is not a wrench certificate is either the same log
            # printed rather than saved, or a map. Try the log first: the map
            # has no header row, so it simply yields nothing.
            try:
                rows = _parse_pdf_log(doc["path"])
            except Exception:                    # noqa: BLE001
                rows = []

        if rows:
            for r in rows:
                flanges.append((
                    project_id, doc["id"], doc["fingerprint"], doc["segment"] or "",
                    r["sheet"], r["row_no"], r["flange_no"], r["nps"],
                    r["pressure_class"], r["gasket"], r["bolts"], r["lubricant"],
                    r["round1"], r["round2"], r["round3"], r["round4"],
                    r["pattern"], r["wrench"], r["wrench_key"], r["cert_checked"],
                    r["inspector"], r["bolted_on"], r["drawing_no"], r["notes"],
                    r["line_size"], r["service"], r["job_start"], SOURCE,
                ))
            continue

        # A PDF here is either a torque wrench certificate or a flange map.
        instrument = parse_bare_serial(doc["filename"])
        if instrument.serial:
            wrenches.append((
                project_id, doc["id"], instrument.kind or "torque_wrench",
                instrument.serial, instrument.serial_key, instrument.calibrated,
                instrument.description, "filename", WRENCH_SOURCE,
            ))
            continue

        try:
            balloons, drawings = _parse_map(doc["path"])
        except Exception:                        # noqa: BLE001
            continue
        if balloons:
            maps.append((project_id, doc["id"], doc["segment"] or "",
                         ", ".join(drawings), balloons, MAP_SOURCE))

    with db.tx() as c:
        c.execute("DELETE FROM flange WHERE project_id=? AND source=?",
                  (project_id, SOURCE))
        c.execute("DELETE FROM flange_map WHERE project_id=? AND source=?",
                  (project_id, MAP_SOURCE))
        c.execute("DELETE FROM instrument_cal WHERE project_id=? AND source=?",
                  (project_id, WRENCH_SOURCE))
        if flanges:
            c.executemany(
                """INSERT INTO flange
                   (project_id, document_id, fingerprint, segment, sheet, row_no,
                    flange_no, nps, pressure_class, gasket, bolts, lubricant,
                    round1, round2, round3, round4, pattern, wrench, wrench_key,
                    cert_checked, inspector, bolted_on, drawing_no, notes,
                    line_size, service, job_start, source)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                flanges,
            )
        if maps:
            c.executemany(
                """INSERT INTO flange_map
                   (project_id, document_id, segment, drawings, balloons, source)
                   VALUES (?,?,?,?,?,?)""",
                maps,
            )
        if wrenches:
            c.executemany(
                """INSERT INTO instrument_cal
                   (project_id, document_id, kind, serial, serial_key, calibrated,
                    description, evidence, source)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                wrenches,
            )
    return len(flanges), len(maps), len(wrenches)
