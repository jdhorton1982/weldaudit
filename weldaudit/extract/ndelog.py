"""NDE rig log - who was allowed to shoot, and from when.

The rig log lists each NDE technician on the job with their company, rig
letter, cert status, visual acuity and cert date.  It is the reference the
qualification rules check reader sheets against.
"""

from __future__ import annotations

import re
import warnings
from datetime import date, datetime
from typing import Any

import openpyxl

from ..db import Database

warnings.filterwarnings("ignore", module="openpyxl")

_COLUMNS: dict[str, tuple[str, ...]] = {
    "company":    ("nde company name", "company", "nde company"),
    "name":       ("tech name", "technician", "name"),
    "rig_letter": ("rig letter", "rig"),
    "certs":      ("certs  yes/no", "certs yes/no", "certs"),
    "acuity":     ("visual acuity  yes/no", "visual acuity yes/no", "visual acuity"),
    "cert_date":  ("cert date",),
    "arrived":    ("date arrived on job", "date arrived"),
}


def _clean(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return re.sub(r"\s+", " ", str(v)).strip()


def _label(v: Any) -> str:
    return _clean(v).lower()


def parse_workbook(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: list[dict] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        grid = [list(r) for r in ws.iter_rows(max_row=200, max_col=40, values_only=True)]
        header_row, colmap = _find_header(grid)
        if header_row is None:
            continue
        blanks = 0
        for row in grid[header_row + 1 :]:
            rec = {f: _clean(row[i]) if i < len(row) else "" for f, i in colmap.items()}
            if not rec.get("name") and not rec.get("company"):
                blanks += 1
                if blanks >= 8:
                    break
                continue
            blanks = 0
            out.append(rec)
    wb.close()
    return out


def _find_header(grid: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
    for r, row in enumerate(grid):
        labels = {i: _label(c) for i, c in enumerate(row) if _label(c)}
        if not any(l in _COLUMNS["name"] for l in labels.values()):
            continue
        colmap: dict[str, int] = {}
        for field, variants in _COLUMNS.items():
            for i, label in labels.items():
                if label in variants and field not in colmap:
                    colmap[field] = i
        if "name" in colmap:
            return r, colmap
    return None, {}


def extract(db: Database, project_id: int) -> tuple[int, int]:
    docs = db.q(
        """SELECT id, path, segment FROM document
           WHERE project_id=? AND kind='nde_rig_log' AND ext IN ('.xlsx','.xlsm')""",
        (project_id,),
    )
    with db.tx() as c:
        c.execute("DELETE FROM nde_tech WHERE project_id=?", (project_id,))

    rows: list[tuple] = []
    ok = 0
    for d in docs:
        try:
            techs = parse_workbook(d["path"])
        except Exception:
            continue
        if not techs:
            continue
        ok += 1
        for t in techs:
            rows.append(
                (
                    project_id, d["segment"], t.get("company", ""), t.get("name", ""),
                    t.get("rig_letter", ""), t.get("certs", ""), t.get("acuity", ""),
                    t.get("cert_date", ""), t.get("arrived", ""),
                )
            )

    if rows:
        with db.tx() as c:
            c.executemany(
                """INSERT INTO nde_tech
                   (project_id, segment, company, name, rig_letter, certs,
                    acuity, cert_date, arrived)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                rows,
            )
    return ok, len(rows)
