"""Daily Weld Reports - the weld map for the paper-based projects.

A DWR is an Excel sheet with a header block (AFE, job, line, size, wall) and a
grid of welds.  The NOTES column carries the NDE report the weld was shot
under, written loosely: ``DTI-5`` for what the reader sheet files as
``DTI-005``.  Column positions drift between segments, so columns are located
by their header label rather than by index.
"""

from __future__ import annotations

import re
import warnings
from datetime import date, datetime
from typing import Any

import openpyxl

from ..db import Database
from ..ids import parse_one

warnings.filterwarnings("ignore", module="openpyxl")

#: Header labels -> the weld column they identify.
_COLUMNS: dict[str, tuple[str, ...]] = {
    "weld_no":     ("weld #", "weld no", "weld number", "weld#"),
    "weld_size":   ("size", "weld size"),
    "weld_type":   ("weld type", "type"),
    "process":     ("process",),
    "welder_root": ("root",),
    "welder_hp":   ("hp", "hot pass"),
    "welder_fill": ("fill",),
    "welder_cap":  ("cap",),
    "notes":       ("notes", "note", "nde"),
}

#: Header-block labels -> the metadata field they carry.
_META: dict[str, tuple[str, ...]] = {
    "afe":         ("afe#", "afe #", "afe"),
    "unit":        ("unit:",),
    "job_name":    ("job name",),
    "contractor":  ("contractor",),
    "advisor":     ("construction advis", "const advisor", "xto construction"),
    "engineer":    ("project engineer",),
    "inspector":   ("job lead inspector", "lead inspector"),
    "report_date": ("today's date", "todays date", "date:"),
    "line_size":   ("line size",),
    "wall":        ("wt",),
    "material":    ("material",),
    "yield":       ("yield",),
    "service":     ("service",),
}


def _clean(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    return re.sub(r"\s+", " ", str(v)).strip()


def _norm_label(v: Any) -> str:
    return _clean(v).lower().rstrip(":").strip()


def parse_workbook(path: str, known_prefixes: set[str] | None = None) -> dict:
    """Read one DWR into ``{"meta": {...}, "welds": [ {...}, ... ]}``."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = [list(r) for r in ws.iter_rows(max_row=200, max_col=60, values_only=True)]
    wb.close()

    meta = _read_meta(grid)
    header_row, colmap = _find_header(grid)
    welds: list[dict] = []
    if header_row is not None:
        welds = _read_welds(grid, header_row, colmap)
        _link_nde(welds, known_prefixes)
    return {"meta": meta, "welds": welds}


def _read_meta(grid: list[list[Any]]) -> dict[str, str]:
    """Label/value pairs from the header block: value is the next filled cell right."""
    meta: dict[str, str] = {}
    for row in grid[:16]:
        for i, cell in enumerate(row):
            label = _norm_label(cell)
            if not label:
                continue
            for field, variants in _META.items():
                if field in meta:
                    continue
                if any(label.startswith(v) for v in variants):
                    for j in range(i + 1, min(i + 12, len(row))):
                        val = _clean(row[j])
                        if val and _norm_label(row[j]) not in _flat_labels():
                            meta[field] = val
                            break
    return meta


_FLAT: set[str] | None = None


def _flat_labels() -> set[str]:
    global _FLAT
    if _FLAT is None:
        _FLAT = {v for variants in _META.values() for v in variants}
    return _FLAT


def _find_header(grid: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
    """Locate the weld grid header and map each wanted column to its index."""
    for r, row in enumerate(grid):
        labels = {i: _norm_label(c) for i, c in enumerate(row) if _norm_label(c)}
        if not any(l in _COLUMNS["weld_no"] for l in labels.values()):
            continue
        colmap: dict[str, int] = {}
        for field, variants in _COLUMNS.items():
            for i, label in labels.items():
                if label in variants and field not in colmap:
                    colmap[field] = i
        if "weld_no" in colmap:
            return r, colmap
    return None, {}


def known_nde_prefixes(db: Database, project_id: int) -> set[str]:
    """The NDE series that actually exist in this project's reader sheets.

    The NOTES column is a free-text field and crews put all sorts in it -
    ``BORE 2``, drawing numbers like ``PG-327``, spool references like
    ``FG-515``.  Treating every ``XX-nnn`` in there as a cited NDE report
    manufactures critical findings for shots that were never NDE at all.  The
    reader sheets themselves are the authority on which prefixes are real.
    """
    return {
        r["prefix"]
        for r in db.q(
            "SELECT DISTINCT prefix FROM nde_shot WHERE project_id=? AND prefix<>''",
            (project_id,),
        )
    }


def _read_welds(grid: list[list[Any]], header_row: int, colmap: dict[str, int]) -> list[dict]:
    welds: list[dict] = []
    blanks = 0
    for row in grid[header_row + 1 :]:
        weld_no = _clean(row[colmap["weld_no"]]) if colmap["weld_no"] < len(row) else ""
        if not weld_no:
            blanks += 1
            if blanks >= 12:      # past the end of the grid
                break
            continue
        blanks = 0
        if not re.match(r"^[\w\-./]+$", weld_no) or len(weld_no) > 24:
            continue              # a stray note, not a weld number

        rec: dict[str, str] = {"weld_no": weld_no}
        for field, idx in colmap.items():
            if field == "weld_no" or idx >= len(row):
                continue
            rec[field] = _clean(row[idx])
        welds.append(rec)
    return welds


def _link_nde(welds: list[dict], known: set[str] | None) -> None:
    """Resolve each weld's NOTES text to an NDE id, in place.

    With no known prefixes, nothing is accepted. That is the case where we know
    *least* about this job's numbering, not most: treating an empty vocabulary
    as "accept anything" turns a spool number like ``DTD22MP-LP-16-1B`` into a
    citation of report ``LP-016``, and then into findings about a weld that
    does not exist. The note is kept either way, and NDE-09 reports what could
    not be resolved.
    """
    for rec in welds:
        note = rec.get("notes", "")
        nid = parse_one(note)
        if nid and known and nid.prefix in known:
            rec["nde_id"] = str(nid)
            rec["note_kind"] = "nde"
        else:
            rec["nde_id"] = ""
            rec["note_kind"] = "other" if note else ""


def extract(db: Database, project_id: int) -> tuple[int, int]:
    """Load every DWR spreadsheet in the project.  Returns ``(files, welds)``.

    Must run *after* the reader sheets, which supply the set of NDE prefixes
    that are real on this project.
    """
    docs = db.q(
        """SELECT id, path, segment FROM document
           WHERE project_id=? AND kind='daily_weld_report'
             AND ext IN ('.xlsx', '.xlsm')""",
        (project_id,),
    )
    known = known_nde_prefixes(db, project_id) or None

    with db.tx() as c:
        c.execute(
            "DELETE FROM weld WHERE project_id=? AND source='daily_weld_report'",
            (project_id,),
        )

    rows: list[tuple] = []
    ok_files = 0
    for d in docs:
        try:
            parsed = parse_workbook(d["path"], known)
        except Exception:
            continue          # unreadable / OneDrive placeholder; surfaced by the rule
        if not parsed["welds"]:
            continue
        ok_files += 1
        meta = parsed["meta"]
        line = meta.get("service") or meta.get("job_name") or ""
        for w in parsed["welds"]:
            rows.append(
                (
                    project_id, d["id"], d["segment"], line,
                    w.get("weld_no", ""), w.get("weld_size", "") or meta.get("line_size", ""),
                    w.get("weld_type", ""), w.get("process", ""),
                    w.get("welder_root", ""), w.get("welder_hp", ""),
                    w.get("welder_fill", ""), w.get("welder_cap", ""),
                    meta.get("report_date", ""), w.get("nde_id", ""),
                    w.get("notes", ""), "daily_weld_report",
                )
            )

    if rows:
        with db.tx() as c:
            c.executemany(
                """INSERT INTO weld
                   (project_id, document_id, segment, line, weld_no, weld_size,
                    weld_type, process, welder_root, welder_hp, welder_fill,
                    welder_cap, date_welded, nde_id, note, source)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                rows,
            )
    return ok_files, len(rows)
