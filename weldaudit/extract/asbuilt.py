"""Loading the as-built sheets — a third weld register, with stationing.

Fourteen workbooks across three jobs, 1,976 joints, and almost every one
carries the NDE report on the weld at its end. That makes the as-built the
largest single register in the corpus, and the only one that says *where* on
the line each joint sits.
"""

from __future__ import annotations

import warnings

import openpyxl

from ..asbuilt import parse_sheet
from ..db import Database
from ..mtrname import normalise_heat
from ..weldmap import format_id, parse_id_token

warnings.filterwarnings("ignore", module="openpyxl")

SOURCE = "asbuilt_xlsx"


def _nde_id(raw: str) -> str:
    """The X-RAY # column normalised, or blank.

    A fifth of the column holds bare row numbers rather than report numbers.
    The isometric grammar rejects those, which is what keeps them from
    joining to anything.
    """
    parsed = parse_id_token(raw)
    return format_id(*parsed) if parsed else ""


def extract(db: Database, project_id: int) -> tuple[int, int]:
    """Load every as-built workbook. Returns ``(joints, documents)``."""
    documents = db.q(
        """SELECT id, path, filename, segment, fingerprint FROM document
           WHERE project_id=? AND (kind='as_built' OR section_no=3)
             AND ext IN ('.xlsx','.xlsm')""",
        (project_id,),
    )

    records: list[tuple] = []
    seen: set[str] = set()
    for doc in documents:
        key = doc["fingerprint"] or f"doc:{doc['id']}"
        if key in seen:
            continue
        seen.add(key)
        try:
            book = openpyxl.load_workbook(doc["path"], data_only=True, read_only=True)
        except Exception:                     # noqa: BLE001 - a bad workbook is
            continue                          # not worth failing the run over

        for name in book.sheetnames:
            try:
                rows = [list(r) for r in book[name].iter_rows(values_only=True)]
            except Exception:                 # noqa: BLE001
                continue
            sheet = parse_sheet(name, rows)
            for joint in sheet.joints:
                records.append((
                    project_id, doc["id"], doc["fingerprint"], doc["segment"] or "",
                    sheet.name, joint.band, joint.seq, joint.station,
                    joint.station_ft, joint.length, joint.heat,
                    normalise_heat(joint.heat), joint.joint_no, joint.size,
                    joint.description, joint.xray, _nde_id(joint.xray),
                    sheet.pipe_size, sheet.grade, sheet.wall, sheet.service,
                    SOURCE,
                ))

    with db.tx() as c:
        c.execute("DELETE FROM asbuilt_joint WHERE project_id=? AND source=?",
                  (project_id, SOURCE))
        if records:
            c.executemany(
                """INSERT INTO asbuilt_joint
                   (project_id, document_id, fingerprint, segment, sheet, band,
                    seq, station, station_ft, length, heat, heat_key, joint_no,
                    size, description, xray, nde_id, pipe_size, grade, wall,
                    service, source)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                records,
            )
    return len(records), len(seen)
