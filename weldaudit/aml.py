"""The Approved Materials List: loading it, and deciding whether a material is on it.

Two jobs live here, both pure logic so they can be tested without a database.

*Matching a manufacturer.*  A mill is written one way on the AML ("Awaji (ASK)",
"NOV (National Oilwell Varco)", "Norvale Silcotub") and another way on a
purchase record ("AWAJI", "NOV", "NORVALE").  Matching is therefore fuzzy, but
never silently so: a match either clears a high bar, or it is returned as a
candidate for a human to confirm, or it fails.  An audit tool that quietly
accepts a near-miss is worse than one that asks.

*Enforcing a Specific Limit.*  430 of the 1,445 AML rows carry one, and most of
them are size restrictions - "Up to NPS 20", "NPS 8 to 42", "NPS 2 and
smaller".  A mill approved up to NPS 20 supplying 24" pipe is a genuine
non-conformance that reading the manufacturer name alone will never catch.
Limits that are not about size ("Induction bends only", "Brand: Capitol Mfg.")
are surfaced for manual review rather than guessed at.
"""

from __future__ import annotations

import re
import warnings
from dataclasses import dataclass, field
from fractions import Fraction
from pathlib import Path

from rapidfuzz import fuzz, process

warnings.filterwarnings("ignore", module="openpyxl")

#: Score at or above which a manufacturer is accepted outright.
CONFIDENT = 90
#: Score at or above which a manufacturer is offered as a candidate to confirm.
PLAUSIBLE = 72

#: Legal forms only.  Industry words like "steel", "tube" and "forge" are
#: deliberately *kept* - stripping them collapses "American Steel Pipe" to
#: "american" (which then matches any American-anything) while simultaneously
#: reducing "Tube Forging Of Ashton" to something a search for "Tube forg" can
#: no longer find.  Both failure directions were observed on this corpus.
_LEGAL = re.compile(
    r"\b(inc|llc|l\.l\.c|ltd|limited|co|corp|corporation|company|gmbh|srl|"
    r"s\.?p\.?a|sas|bv|b\.v|nv|n\.v|ag|a\.?s|pvt|private|plc|kg|oy|ab)\b\.?",
    re.IGNORECASE,
)

#: A trailing parenthetical or bracketed qualifier: "Awaji (ASK)", "Bebitz (F)".
_PAREN = re.compile(r"[\(\[][^)\]]*[\)\]]")

#: Status markers the AML puts in front of or inside a name.
_STATUS = re.compile(r"^\s*\*+\s*|\s*[-–]\s*(HOLD|PROVISIONAL|PROV)\b.*$", re.IGNORECASE)

#: An entry that has been withdrawn or replaced must not silently approve.
_SUPERSEDED = re.compile(r"\(deleted\)|\b(updated|changed|moved|renamed)\s+to\b", re.IGNORECASE)


def normalise_manufacturer(name: str) -> str:
    """Reduce a manufacturer name to its identifying core, lower case."""
    text = _STATUS.sub(" ", name or "")
    text = _PAREN.sub(" ", text)
    # Strip legal forms before punctuation goes, so dotted variants like
    # "S.p.A." are recognised, and again afterwards for the undotted "Inc".
    text = _LEGAL.sub(" ", text)
    text = re.sub(r"[^\w\s&-]", " ", text)
    text = _LEGAL.sub(" ", text)
    text = re.sub(r"[\s\-_]+", " ", text).strip().lower()
    return text


def _token_prefix_match(query: str, candidate: str) -> bool:
    """Every query token is a whole-word prefix of a candidate token, in order.

    Lets "tube forg" find "Tube Forging Of Ashton" without letting a
    three-letter query match anything that merely contains those letters.
    """
    q, c = query.split(), candidate.split()
    if not q or len(q) > len(c):
        return False
    i = 0
    for token in c:
        if token.startswith(q[i]):
            i += 1
            if i == len(q):
                return True
    return False


# ---------------------------------------------------------------------------
# Nominal pipe size
# ---------------------------------------------------------------------------

_FRACTION = {
    "½": ".5", "¼": ".25", "¾": ".75", "⅜": ".375", "⅝": ".625", "⅞": ".875", "⅛": ".125",
}


def parse_nps(text: str | float | int | None) -> float | None:
    """Nominal pipe size in inches from the many ways it gets written.

    Handles ``16"``, ``16”``, ``NPS 1½``, ``1-1/2``, ``3/4``, ``4 IN``, ``4in``.
    """
    if text is None:
        return None
    if isinstance(text, (int, float)):
        return float(text) or None

    s = str(text).strip()
    for glyph, value in _FRACTION.items():
        s = s.replace(glyph, value)
    s = s.replace("”", '"').replace("″", '"')
    s = re.sub(r"\b(nps|dn|sch(edule)?)\b", " ", s, flags=re.IGNORECASE)

    # "1-1/2" or "1 1/2"
    m = re.search(r"(\d+)\s*[-\s]\s*(\d+)\s*/\s*(\d+)", s)
    if m:
        whole, num, den = (int(g) for g in m.groups())
        return whole + num / den if den else float(whole)
    # bare fraction "3/4"
    m = re.search(r"(?<![\d.])(\d+)\s*/\s*(\d+)", s)
    if m:
        num, den = int(m.group(1)), int(m.group(2))
        return num / den if den else None
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m:
        return None
    value = float(m.group(1))
    # Guard against picking up a wall thickness or a pressure class.
    return value if 0.125 <= value <= 96 else None


# ---------------------------------------------------------------------------
# Specific Limits
# ---------------------------------------------------------------------------


@dataclass
class SizeLimit:
    min_nps: float | None = None
    max_nps: float | None = None

    def allows(self, nps: float) -> bool:
        if self.min_nps is not None and nps < self.min_nps - 1e-9:
            return False
        if self.max_nps is not None and nps > self.max_nps + 1e-9:
            return False
        return True

    def describe(self) -> str:
        if self.min_nps is not None and self.max_nps is not None:
            return f"NPS {_n(self.min_nps)} to {_n(self.max_nps)}"
        if self.max_nps is not None:
            return f"up to NPS {_n(self.max_nps)}"
        if self.min_nps is not None:
            return f"NPS {_n(self.min_nps)} and larger"
        return "no size limit"


def _n(v: float) -> str:
    """Render a pipe size the way the trade writes it: 1.5 -> ``1-1/2``."""
    if float(v).is_integer():
        return str(int(v))
    frac = Fraction(v).limit_denominator(16)
    whole, rem = divmod(frac.numerator, frac.denominator)
    return f"{whole}-{rem}/{frac.denominator}" if whole else f"{rem}/{frac.denominator}"


#: "Up to NPS 6".  The NPS/inch marker is required: "Up to 9% Cr, Up to NPS 6"
#: must yield 6, not 9.
_UP_TO = re.compile(
    r"\bup to\s+(?:and including\s+)?(?:nps|dn)\s*([\d./\s½¼¾⅜⅝⅞⅛-]+)", re.IGNORECASE
)
_RANGE = re.compile(r"\bnps\s*([\d./\s½¼¾⅜⅝⅞⅛-]+?)\s*(?:to|through|thru|-)\s*(?:nps\s*)?([\d./½¼¾⅜⅝⅞⅛]+)", re.IGNORECASE)
_AND_LARGER = re.compile(r"\bnps\s*([\d./\s½¼¾⅜⅝⅞⅛-]+?)\s*and\s+(?:larger|greater|above)", re.IGNORECASE)
_AND_SMALLER = re.compile(r"\bnps\s*([\d./\s½¼¾⅜⅝⅞⅛-]+?)\s*and\s+(?:smaller|less|below)", re.IGNORECASE)


def parse_limit(text: str | None) -> tuple[SizeLimit | None, str]:
    """Split a Specific Limit into an enforceable size rule and leftover conditions.

    Returns ``(size_limit_or_None, conditions_text)``.  ``conditions_text`` is
    whatever could not be turned into a size rule and therefore needs a human -
    "Induction bends only", "Brand: Truseal", "Assembly and testing only".
    """
    raw = (text or "").strip()
    if not raw or raw in ("0", "-"):
        return None, ""

    limit: SizeLimit | None = None
    consumed: list[tuple[int, int]] = []

    def take(m: re.Match) -> None:
        consumed.append(m.span())

    if m := _RANGE.search(raw):
        lo, hi = parse_nps(m.group(1)), parse_nps(m.group(2))
        if lo is not None and hi is not None:
            limit = SizeLimit(min_nps=lo, max_nps=hi)
            take(m)
    if limit is None and (m := _UP_TO.search(raw)):
        hi = parse_nps(m.group(1))
        if hi is not None:
            limit = SizeLimit(max_nps=hi)
            take(m)
    if limit is None and (m := _AND_LARGER.search(raw)):
        lo = parse_nps(m.group(1))
        if lo is not None:
            limit = SizeLimit(min_nps=lo)
            take(m)
    if limit is None and (m := _AND_SMALLER.search(raw)):
        hi = parse_nps(m.group(1))
        if hi is not None:
            limit = SizeLimit(max_nps=hi)
            take(m)

    rest = raw
    for start, end in sorted(consumed, reverse=True):
        rest = rest[:start] + " " + rest[end:]
    rest = re.sub(r"^[\s,.;]+|[\s,.;]+$", "", re.sub(r"\s{2,}", " ", rest))
    return limit, rest


# ---------------------------------------------------------------------------
# Entries and matching
# ---------------------------------------------------------------------------


@dataclass
class AmlEntry:
    category: str            # the AML sheet, e.g. "1.0 Pipe"
    manufacturer: str
    location: str
    limits_raw: str
    size_limit: SizeLimit | None = None
    conditions: str = ""
    key: str = field(default="", repr=False)

    @property
    def on_hold(self) -> bool:
        return bool(re.search(r"\bHOLD\b|sanction", self.manufacturer + self.limits_raw, re.IGNORECASE))

    @property
    def provisional(self) -> bool:
        return bool(re.search(r"PROVISIONAL|\bPROV\b|^\s*\*", self.manufacturer, re.IGNORECASE))

    @property
    def superseded(self) -> bool:
        """Marked ``(deleted)`` or redirected to another entry."""
        return bool(_SUPERSEDED.search(self.manufacturer))

    @property
    def usable(self) -> bool:
        return not (self.superseded or self.on_hold)

    def caveat(self) -> str:
        notes = []
        if self.superseded:
            notes.append("AML entry is marked deleted or superseded")
        if self.on_hold:
            notes.append("AML entry is on HOLD")
        if self.provisional:
            notes.append("AML entry is provisional")
        return "; ".join(notes)


@dataclass
class MatchResult:
    status: str                     # 'approved' | 'confirm' | 'not_listed'
    entries: list[AmlEntry] = field(default_factory=list)
    score: int = 0
    reason: str = ""


#: AML sheet -> the kinds of material it governs.  Used to avoid clearing a
#: flange against the pipe list, which would let a real exception through.
CATEGORY_KEYWORDS: dict[str, tuple[str, ...]] = {
    "1.0 Pipe": ("pipe", "nipple"),
    "2.0 Pipe Fittings": ("fitting", "elbow", "el ", "tee", "reducer", "cap", "olet",
                          "weldolet", "sockolet", "thredolet", "flexolet", "bend", "stub"),
    "3.0 Flanges": ("flange", "flg", "rfwn", "ffwn", "blind", "wn ", "slip on", "so "),
    "4.0 Pipe Gaskets": ("gasket", "ring joint", "spiral wound"),
    "5.0 Valves": ("valve", "gate", "globe", "check"),
    "6.0 Valves": ("butterfly",),
    "7.0 Valves - Ball": ("ball valve", "trunnion", "bv "),
    "8.0 Valves - Plug": ("plug valve",),
    "9.0 Valves - Special": ("hf acid",),
    "10.0 Welding Consumables": ("electrode", "filler", "consumable"),
    "11.0 Fasteners": ("stud", "bolt", "nut", "fastener"),
    "13.0 Pipe Unions": ("union",),
}


def categories_for(description: str) -> list[str]:
    """Which AML sheets could govern a material with this description."""
    text = f" {(description or '').lower()} "
    hits = [
        sheet for sheet, words in CATEGORY_KEYWORDS.items()
        if any(w in text for w in words)
    ]
    return hits


class Aml:
    """The approved materials list, ready to be queried."""

    def __init__(self, entries: list[AmlEntry], aliases: dict[str, str] | None = None):
        #: Names the certificates use for a company the list calls something
        #: else. Passed in rather than loaded here so a test builds a matcher
        #: that depends on nothing outside itself.
        self._aliases = aliases or {}
        self.entries = entries
        self._by_key: dict[str, list[AmlEntry]] = {}
        for e in entries:
            self._by_key.setdefault(e.key, []).append(e)
        self._keys = list(self._by_key)

    def __len__(self) -> int:
        return len(self.entries)

    # -- loading ------------------------------------------------------------

    @classmethod
    def from_workbook(cls, path: str | Path) -> "Aml":
        """Read an ``AML Search Spreadsheet.xlsx``.

        The flattened ``AllData`` sheet is preferred; the per-category sheets
        are used as a fallback so the loader survives a re-exported workbook.
        """
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        rows: list[tuple[str, str, str, str]] = []

        if "AllData" in wb.sheetnames:
            ws = wb["AllData"]
            for r in ws.iter_rows(min_row=2, max_col=4, values_only=True):
                sheet, mfr, loc, lim = (str(v).strip() if v is not None else "" for v in r)
                rows.append((sheet, mfr, loc, lim))
        else:
            for sheet in wb.sheetnames:
                if not re.match(r"\d", sheet):
                    continue
                ws = wb[sheet]
                for r in ws.iter_rows(min_row=2, max_col=4, values_only=True):
                    _num, mfr, loc, lim = (str(v).strip() if v is not None else "" for v in r)
                    rows.append((sheet, mfr, loc, lim))
        wb.close()

        entries: list[AmlEntry] = []
        for sheet, mfr, loc, lim in rows:
            # Rows without a location are section headings ("Carbon Steel - Seamless").
            if not mfr or not loc:
                continue
            size_limit, conditions = parse_limit(lim)
            entries.append(
                AmlEntry(
                    category=sheet, manufacturer=mfr, location=loc,
                    limits_raw=lim if lim not in ("0",) else "",
                    size_limit=size_limit, conditions=conditions,
                    key=normalise_manufacturer(mfr),
                )
            )
        return cls(entries)

    # -- querying -----------------------------------------------------------

    def _resolve(self, manufacturer: str) -> str:
        """Swap a certificate's name for the AML's, where they differ."""
        return self._aliases.get(normalise_manufacturer(manufacturer), manufacturer)

    def match(self, manufacturer: str, categories: list[str] | None = None) -> MatchResult:
        """Look a manufacturer up, optionally restricted to certain AML sheets."""
        manufacturer = self._resolve(manufacturer)
        key = normalise_manufacturer(manufacturer)
        if not key:
            return MatchResult("not_listed", reason="no manufacturer recorded")

        def keep(entries: list[AmlEntry]) -> list[AmlEntry]:
            if not categories:
                return entries
            narrowed = [e for e in entries if e.category in categories]
            return narrowed or entries

        if key in self._by_key:
            return self._settle(keep(self._by_key[key]), 100, "exact name match")

        # A short trade name is often a whole-word prefix of the full AML name:
        # "norvale" -> "Norvale Silcotub", "tube forg" -> "Tube Forging Of Ashton".
        prefixed = [k for k in self._keys if _token_prefix_match(key, k) or _token_prefix_match(k, key)]
        if prefixed:
            # Several distinct companies can share a leading word ("Norvale
            # Dalmine", "Norvale Algoma"); that is one company's mills, but
            # "Liberty Pipes" and "Liberty Forge" are not. Let _settle decide.
            entries = [e for k in prefixed for e in self._by_key[k]]
            return self._settle(keep(entries), 95, "matched on a leading-word prefix")

        hits = process.extract(key, self._keys, scorer=fuzz.token_set_ratio, limit=5)
        if not hits:
            return MatchResult("not_listed", reason="no similar name on the AML")

        best_key, score, _ = hits[0]
        if score >= CONFIDENT:
            return self._settle(keep(self._by_key[best_key]), int(score), "close name match")
        if score >= PLAUSIBLE:
            candidates = [e for k, s, _ in hits if s >= PLAUSIBLE for e in self._by_key[k][:1]]
            return MatchResult("confirm", candidates, int(score),
                               "close to an AML entry but not an exact match")
        return MatchResult("not_listed", self._by_key[best_key][:1], int(score),
                           "no AML entry with a similar name")

    def nearest(self, name: str) -> tuple[int, str]:
        """Closest AML name by direct similarity, ignoring the prefix rule.

        ``match`` is generous on purpose: a buyer typing "norvale" should reach
        "Norvale Silcotub", so a leading-word prefix scores 95. That generosity
        is wrong when the question is *which line of a page is the
        manufacturer* — there, "MEGA" reaches a shipping address reading
        "MEGAPAD TAKEAWAY" and "Component" reaches "Forged Components Inc".
        This is the strict comparison for callers choosing between candidates.
        """
        key = normalise_manufacturer(self._resolve(name))
        if not key or not self._keys:
            return 0, ""
        best = process.extractOne(key, self._keys, scorer=fuzz.token_set_ratio)
        if not best:
            return 0, ""
        matched, score, _ = best
        entries = self._by_key.get(matched) or []
        return int(score), entries[0].manufacturer if entries else ""

    def _settle(self, entries: list[AmlEntry], score: int, reason: str) -> MatchResult:
        """Turn candidate entries into a verdict.

        An entry that is deleted, superseded or on HOLD cannot approve anything
        on its own - if those are all we found, the answer is "confirm", not
        "approved".  Likewise when the name matched several genuinely different
        companies: a human picks, this tool does not.
        """
        if not entries:
            return MatchResult("not_listed", reason=reason)

        usable = [e for e in entries if e.usable]
        if not usable:
            return MatchResult(
                "confirm", entries, score,
                entries[0].caveat() or "only withdrawn AML entries match this name",
            )

        distinct = {normalise_manufacturer(e.manufacturer) for e in usable}
        if len(distinct) > 1:
            return MatchResult(
                "confirm", usable, score,
                f"name matches {len(distinct)} different AML manufacturers",
            )
        return MatchResult("approved", usable, score, reason)

    def check_size(self, entries: list[AmlEntry], nps: float | None) -> tuple[list[AmlEntry], list[AmlEntry]]:
        """Split candidate entries into ``(allowing, forbidding)`` for a size."""
        if nps is None:
            return list(entries), []
        allowing = [e for e in entries if e.size_limit is None or e.size_limit.allows(nps)]
        forbidding = [e for e in entries if e.size_limit is not None and not e.size_limit.allows(nps)]
        return allowing, forbidding
